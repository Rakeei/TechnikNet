from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from .models import Property, Team, TeamMember, PropertyImage
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime

def get_user_teams(user):
    """Get all teams the user belongs to"""
    if user.is_superuser:
        return Team.objects.all()
    return Team.objects.filter(members__user=user)

def get_user_properties(user):
    """Get all properties accessible by the user"""
    if user.is_superuser:
        return Property.objects.all()
    user_teams = get_user_teams(user)
    return Property.objects.filter(teams__in=user_teams).distinct()
@login_required
def property_list(request):
    # ⭐ Debug: چک کردن user
    print(f"="*50)
    print(f"🔍 User: {request.user.username}")
    print(f"🔍 Is superuser: {request.user.is_superuser}")
    
    properties = get_user_properties(request.user)
    print(f"🔍 Properties from get_user_properties: {properties.count()}")
    
    # Exclude completed properties from main list
    properties = properties.exclude(status__in=['ausbau_abgeschlossen', 'bezahlt'])
    print(f"🔍 After exclude completed: {properties.count()}")
    properties = get_user_properties(request.user)
    
    # Exclude completed properties from main list
    properties = properties.exclude(status__in=['ausbau_abgeschlossen', 'bezahlt'])
    
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    team_filter = request.GET.get('team', '')
    
    if search:
        properties = properties.filter(
            Q(number__icontains=search) |
            Q(village__icontains=search) |
            Q(owner_name__icontains=search) |
            Q(owner_surname__icontains=search) |
            Q(pop_code__icontains=search)
        )
    
    if status_filter:
        properties = properties.filter(status=status_filter)
    
    if team_filter:
        properties = properties.filter(teams__id=team_filter)
    
    # Sort by HBG=Ja first, then by nearest ausbau_termin
    from django.db.models import Case, When, Value, IntegerField
    properties = properties.annotate(
        hbg_priority=Case(
            When(hbg='Ja', then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    ).order_by('hbg_priority', 'ausbau_termin')
    
    paginator = Paginator(properties, 50)
    page = request.GET.get('page', 1)
    properties_page = paginator.get_page(page)
    
    user_teams = get_user_teams(request.user)
    
    # ⭐ Fix: Get statuses correctly for superuser
    if request.user.is_superuser:
        statuses = Property.objects.exclude(
            status__in=['ausbau_abgeschlossen', 'bezahlt']
        ).values_list('status', flat=True).distinct()
    else:
        statuses = Property.objects.filter(
            teams__in=user_teams
        ).exclude(
            status__in=['ausbau_abgeschlossen', 'bezahlt']
        ).values_list('status', flat=True).distinct()
    
    context = {
        'properties': properties_page,
        'teams': user_teams,
        'statuses': statuses,
        'search': search,
        'status_filter': status_filter,
        'team_filter': team_filter,
    }
    return render(request, 'properties/property_list.html', context)

@login_required
def property_detail(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    
    # Check access
    if not request.user.is_superuser:
        user_teams = get_user_teams(request.user)
        if not property_obj.teams.filter(id__in=user_teams).exists():
            messages.error(request, 'Access denied: You do not have permission to view this property')
            return redirect('property_list')
    
    context = {
        'property': property_obj,
        'images': property_obj.images.all()
    }
    return render(request, 'properties/property_detail.html', context)

@login_required
def property_create(request):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')
    
    if request.method == 'POST':
        try:
            property_obj = Property.objects.create(
                number=request.POST.get('number'),
                address_id=request.POST.get('address_id', ''),
                village=request.POST.get('village', ''),
                street=request.POST.get('street', ''),
                house_number=request.POST.get('house_number', ''),
                house_number_affix=request.POST.get('house_number_affix', ''),
                owner_email=request.POST.get('owner_email', ''),
                owner_name=request.POST.get('owner_name', ''),
                owner_surname=request.POST.get('owner_surname', ''),
                owner_phone_1=request.POST.get('owner_phone_1', ''),
                owner_phone_2=request.POST.get('owner_phone_2', ''),
                pop_code=request.POST.get('pop_code', ''),
            )
            
            # Assign teams
            team_ids = request.POST.getlist('teams')
            if team_ids:
                property_obj.teams.set(team_ids)
            
            images = request.FILES.getlist('images')
            for image in images:
                PropertyImage.objects.create(
                    property=property_obj,
                    image=image,
                    uploaded_by=request.user
                )
            
            messages.success(request, f'Property {property_obj.number} created successfully')
            return redirect('property_detail', pk=property_obj.pk)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    teams = Team.objects.all()
    return render(request, 'properties/property_form.html', {'teams': teams, 'mode': 'create'})

@login_required
def property_admin_edit(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')
    
    property_obj = get_object_or_404(Property, pk=pk)
    
    if request.method == 'POST':
        try:
            property_obj.number = request.POST.get('number')
            property_obj.address_id = request.POST.get('address_id', '')
            property_obj.village = request.POST.get('village', '')
            property_obj.street = request.POST.get('street', '')
            property_obj.house_number = request.POST.get('house_number', '')
            property_obj.house_number_affix = request.POST.get('house_number_affix', '')
            property_obj.owner_email = request.POST.get('owner_email', '')
            property_obj.owner_name = request.POST.get('owner_name', '')
            property_obj.owner_surname = request.POST.get('owner_surname', '')
            property_obj.owner_phone_1 = request.POST.get('owner_phone_1', '')
            property_obj.owner_phone_2 = request.POST.get('owner_phone_2', '')
            property_obj.pop_code = request.POST.get('pop_code', '')
            property_obj.save()
            
            # Update teams
            team_ids = request.POST.getlist('teams')
            property_obj.teams.set(team_ids)
            
            images = request.FILES.getlist('images')
            for image in images:
                PropertyImage.objects.create(
                    property=property_obj,
                    image=image,
                    uploaded_by=request.user
                )
            
            messages.success(request, 'Admin fields updated successfully')
            return redirect('property_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    teams = Team.objects.all()
    return render(request, 'properties/property_form.html', {
        'property': property_obj,
        'teams': teams,
        'mode': 'admin_edit'
    })

@login_required
def property_user_edit(request, pk):
    property_obj = get_object_or_404(Property, pk=pk)
    
    # Check if property can be edited by regular users
    if not property_obj.can_user_edit() and not request.user.is_superuser:
        messages.error(request, 'This property is completed and cannot be edited')
        return redirect('property_detail', pk=pk)
    
    # Check access
    if not request.user.is_superuser:
        user_teams = get_user_teams(request.user)
        if not property_obj.teams.filter(id__in=user_teams).exists():
            messages.error(request, 'Access denied')
            return redirect('property_list')
    
    if request.method == 'POST':
        try:
            # Parse integer fields with proper None handling
            property_obj.gebaute_units = int(request.POST.get('gebaute_units')) if request.POST.get('gebaute_units') else None
            property_obj.kl_15m = int(request.POST.get('kl_15m')) if request.POST.get('kl_15m') else None
            property_obj.kl_20m = int(request.POST.get('kl_20m')) if request.POST.get('kl_20m') else None
            property_obj.kl_30m = int(request.POST.get('kl_30m')) if request.POST.get('kl_30m') else None
            property_obj.kl_50m = int(request.POST.get('kl_50m')) if request.POST.get('kl_50m') else None
            property_obj.kl_80m = int(request.POST.get('kl_80m')) if request.POST.get('kl_80m') else None
            property_obj.kl_100m = int(request.POST.get('kl_100m')) if request.POST.get('kl_100m') else None
            
            # Text fields
            property_obj.hbg = request.POST.get('hbg', '')
            
            # DateTime fields - convert from datetime-local format
            hbg_termin_str = request.POST.get('hbg_termin')
            if hbg_termin_str:
                try:
                    property_obj.hbg_termin = timezone.make_aware(datetime.strptime(hbg_termin_str, '%Y-%m-%dT%H:%M'))
                except:
                    property_obj.hbg_termin = None
            else:
                property_obj.hbg_termin = None
            
            ausbau_termin_str = request.POST.get('ausbau_termin')
            if ausbau_termin_str:
                try:
                    property_obj.ausbau_termin = timezone.make_aware(datetime.strptime(ausbau_termin_str, '%Y-%m-%dT%H:%M'))
                except:
                    property_obj.ausbau_termin = None
            else:
                property_obj.ausbau_termin = None
            
            property_obj.keller = request.POST.get('keller', '')
            property_obj.huep = request.POST.get('huep', '')
            property_obj.spleissen = request.POST.get('spleissen', '')
            property_obj.comments = request.POST.get('comments', '')
            property_obj.ohne_infra = int(request.POST.get('ohne_infra')) if request.POST.get('ohne_infra') else 0
            property_obj.mit_infra = int(request.POST.get('mit_infra')) if request.POST.get('mit_infra') else 0
            property_obj.status = request.POST.get('status', '')
            
            property_obj.save()
                        
            # Handle image uploads
            images = request.FILES.getlist('images')
            if images:
                for image in images:
                    PropertyImage.objects.create(
                        property=property_obj,
                        image=image,
                        uploaded_by=request.user
                    )
                messages.success(request, f'Property updated and {len(images)} image(s) uploaded successfully')
            else:
                messages.success(request, 'User fields updated successfully')
            messages.success(request, 'User fields updated successfully')
            return redirect('property_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'properties/property_user_form.html', {'property': property_obj})

@login_required
def property_delete(request, pk):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')
    
    property_obj = get_object_or_404(Property, pk=pk)
    
    if request.method == 'POST':
        number = property_obj.number
        property_obj.delete()
        messages.success(request, f'Property {number} deleted')
        return redirect('property_list')
    
    return render(request, 'properties/property_confirm_delete.html', {'property': property_obj})

@login_required
def image_delete(request, pk):
    image = get_object_or_404(PropertyImage, pk=pk)
    property_pk = image.property.pk
    
    # Check access
    if not request.user.is_superuser:
        user_teams = get_user_teams(request.user)
        if not image.property.teams.filter(id__in=user_teams).exists():
            messages.error(request, 'Access denied')
            return redirect('property_list')
    
    if request.method == 'POST':
        image.delete()
        messages.success(request, 'Image deleted')
    
    return redirect('property_detail', pk=property_pk)

@login_required
def property_upload_image(request, pk):
    """Quick image upload from property list"""
    property_obj = get_object_or_404(Property, pk=pk)
    
    # Check access
    if not request.user.is_superuser:
        user_teams = get_user_teams(request.user)
        if not property_obj.teams.filter(id__in=user_teams).exists():
            messages.error(request, 'Access denied')
            return redirect('property_list')
    
    if request.method == 'POST':
        images = request.FILES.getlist('images')
        if images:
            count = 0
            for image in images:
                PropertyImage.objects.create(
                    property=property_obj,
                    image=image,
                    uploaded_by=request.user
                )
                count += 1
            messages.success(request, f'{count} image(s) uploaded successfully')
        else:
            messages.warning(request, 'No images selected')
        
        # Redirect back to list or detail based on 'next' parameter
        next_url = request.GET.get('next', 'property_list')
        if next_url == 'detail':
            return redirect('property_detail', pk=pk)
        return redirect('property_list')
    
    context = {
        'property': property_obj,
    }
    return render(request, 'properties/property_upload_image.html', context)

@login_required
def property_completed(request):
    """List of completed properties (Ausbau Abgeschlossen and Bezahlt)"""
    properties = get_user_properties(request.user)
    
    # Only show completed properties
    properties = properties.filter(status__in=['ausbau_abgeschlossen', 'bezahlt'])
    
    search = request.GET.get('search', '').strip()
    team_filter = request.GET.get('team', '')
    status_filter = request.GET.get('status', '')
    
    if search:
        properties = properties.filter(
            Q(number__icontains=search) |
            Q(village__icontains=search) |
            Q(owner_name__icontains=search) |
            Q(owner_surname__icontains=search)
        )
    
    if team_filter:
        properties = properties.filter(teams__id=team_filter)
    
    if status_filter:
        properties = properties.filter(status=status_filter)
    
    properties = properties.order_by('-updated_at')
    
    paginator = Paginator(properties, 50)
    page = request.GET.get('page', 1)
    properties_page = paginator.get_page(page)
    
    user_teams = get_user_teams(request.user)
    
    context = {
        'properties': properties_page,
        'teams': user_teams,
        'search': search,
        'team_filter': team_filter,
        'status_filter': status_filter,
    }
    return render(request, 'properties/property_completed.html', context)

@login_required
def property_completed_edit(request, pk):
    """Admin only - change status from Ausbau Abgeschlossen to Bezahlt"""
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_completed')
    
    property_obj = get_object_or_404(Property, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['ausbau_abgeschlossen', 'bezahlt']:
            property_obj.status = new_status
            property_obj.save()
            messages.success(request, f'Status updated to {property_obj.get_status_display()}')
        else:
            messages.error(request, 'Invalid status')
        return redirect('property_completed')
    
    return redirect('property_completed')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
import io

@login_required
def excel_import_export(request):
    """Admin page for importing and exporting Excel files"""
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')
    
    context = {
        'total_properties': Property.objects.count(),
        'teams': Team.objects.all(),
    }
    return render(request, 'properties/excel_import_export.html', context)
@login_required
def excel_import(request):
    """Handle Excel file import with detailed error reporting"""
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')
    
    if request.method != 'POST':
        return redirect('excel_import_export')
    
    excel_file = request.FILES.get('excel_file')
    force_replace = request.POST.get('force_replace') == 'on'
    default_team_id = request.POST.get('default_team')
    
    if not excel_file:
        messages.error(request, 'Please select an Excel file')
        return redirect('excel_import_export')
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Invalid file format. Please upload .xlsx or .xls file')
        return redirect('excel_import_export')
    
    try:
        import pandas as pd
        from datetime import datetime
        
        # Read Excel file
        df = pd.read_excel(excel_file)
        
        # نمایش ستون‌های موجود در فایل
        available_columns = df.columns.tolist()
        print(f"📊 Available columns in Excel: {available_columns}")
        
        # Clean data - فقط NaN های pandas را جایگزین می‌کنیم
        df = df.fillna('')
        
        # Get default team
        default_team = None
        if default_team_id:
            try:
                default_team = Team.objects.get(id=default_team_id)
            except Team.DoesNotExist:
                pass
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0
        errors = []
        
        def clean_str(value):
            """Convert value to string, return empty string if None/NaN"""
            if value is None or value == '':
                return ''
            if pd.isna(value):
                return ''
            return str(value).strip()
        
        def parse_datetime(date_str):
            """Parse datetime, return None if invalid/empty"""
            if date_str is None or date_str == '':
                return None
            if pd.isna(date_str):
                return None
            try:
                if isinstance(date_str, str):
                    if not date_str.strip():
                        return None
                    dt = datetime.strptime(date_str.split()[0], '%Y-%m-%d')
                elif hasattr(date_str, 'to_pydatetime'):
                    dt = date_str.to_pydatetime()
                else:
                    return None
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                return dt
            except:
                return None
        
        def safe_get(row, column_name, default=''):
            """Safely get column value even if column doesn't exist"""
            if column_name not in df.columns:
                return default
            value = row.get(column_name, default)
            if pd.isna(value) or value == '':
                return default
            return value
        
        def safe_int(value, default=None):
            """Safely convert to integer, return default if empty/invalid"""
            if value is None or value == '':
                return default
            if pd.isna(value):
                return default
            try:
                # Handle string representations
                str_val = str(value).strip().lower()
                if str_val in ['', 'nan', 'none', 'null']:
                    return default
                return int(float(value))
            except:
                return default
        
        for idx, row in df.iterrows():
            row_number = idx + 2  # Excel row number (header is row 1)
            try:
                # Get Number field (required - این تنها فیلد الزامی است)
                number = clean_str(safe_get(row, 'Number', ''))
                if not number:
                    error_count += 1
                    errors.append(f"Row {row_number}: ❌ Missing 'Number' field (required)")
                    continue
                
                # Check if exists
                exists = Property.objects.filter(number=number).exists()
                
                if exists and not force_replace:
                    skipped_count += 1
                    continue
                
                # Prepare property data - همه فیلدها اختیاری هستند
                property_data = {
                    'address_id': clean_str(safe_get(row, 'Address ID', '')),
                    'village': clean_str(safe_get(row, 'Village', '')),
                    'street': clean_str(safe_get(row, 'Street', '')),
                    'house_number': clean_str(safe_get(row, 'House number', '')),
                    'house_number_affix': clean_str(safe_get(row, 'House number affix', '')),
                    'owner_email': clean_str(safe_get(row, 'Owner email', '')),
                    'owner_name': clean_str(safe_get(row, 'Owner name', '')),
                    'owner_surname': clean_str(safe_get(row, 'Owner surname', '')),
                    'owner_phone_1': clean_str(safe_get(row, 'Owner phone 1', '')),
                    'owner_phone_2': clean_str(safe_get(row, 'Owner phone 2', '')),
                    'pop_code': clean_str(safe_get(row, 'PoP code', '')),
                    'gebaute_units': safe_int(safe_get(row, 'Gebaute Units'), None),
                    'hbg': clean_str(safe_get(row, 'HBG', '')),
                    'hbg_termin': parse_datetime(safe_get(row, 'HBG Termin')),
                    'ausbau_termin': parse_datetime(safe_get(row, 'Ausbau Termin')),
                    'kl_15m': safe_int(safe_get(row, 'K.L 15M'), 0),
                    'kl_20m': safe_int(safe_get(row, 'K.L 20M'), 0),
                    'kl_30m': safe_int(safe_get(row, 'K.L 30M'), 0),
                    'kl_50m': safe_int(safe_get(row, 'K.L 50M'), 0),
                    'kl_80m': safe_int(safe_get(row, 'K.L 80M'), 0),
                    'kl_100m': safe_int(safe_get(row, 'K.L 100M'), 0),
                    'keller': clean_str(safe_get(row, 'keller', '')),
                    'huep': clean_str(safe_get(row, 'HÜP', '')),
                    'spleissen': clean_str(safe_get(row, 'spleissen', '')),
                    'ohne_infra': safe_int(safe_get(row, 'ohne Infra'), 0),
                    'mit_infra': safe_int(safe_get(row, 'mit Infra'), 0),
                    'status': clean_str(safe_get(row, 'Status', '')),
                    'comments': clean_str(safe_get(row, 'Comments', '')),
                }
                
                # Create or update property
                property_obj, created = Property.objects.update_or_create(
                    number=number,
                    defaults=property_data
                )
                
                # Assign teams from Excel column or default team
                team_names_str = clean_str(safe_get(row, 'Team', ''))
                teams_assigned = []
                
                if team_names_str:
                    # Clear existing teams first for updates
                    if not created:
                        property_obj.teams.clear()
                    
                    # Split by comma and find teams
                    team_names = [name.strip() for name in team_names_str.split(',') if name.strip()]
                    for team_name in team_names:
                        try:
                            team = Team.objects.get(name=team_name)
                            property_obj.teams.add(team)
                            teams_assigned.append(team_name)
                        except Team.DoesNotExist:
                            errors.append(f"Row {row_number}: ⚠️ Team '{team_name}' not found for property {number}")
                
                elif default_team and created:
                    # Use default team only for new properties
                    property_obj.teams.add(default_team)
                    teams_assigned.append(default_team.name)
                
                # Success
                action = "Created" if created else "Updated"
                team_info = f" → Teams: {', '.join(teams_assigned)}" if teams_assigned else ""
                print(f"✅ Row {row_number}: {action} property '{number}'{team_info}")
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                error_count += 1
                error_msg = f"Row {row_number} (Number: {number if 'number' in locals() else 'Unknown'}): ❌ {str(e)}"
                errors.append(error_msg)
                print(error_msg)
                import traceback
                traceback.print_exc()
        
        # Show detailed results
        if created_count > 0:
            messages.success(request, f'✅ Created {created_count} new properties')
        if updated_count > 0:
            messages.success(request, f'✅ Updated {updated_count} existing properties')
        if skipped_count > 0:
            messages.info(request, f'ℹ️ Skipped {skipped_count} existing properties (check "Force Replace" to update them)')
        if error_count > 0:
            messages.warning(request, f'⚠️ {error_count} rows had errors:')
            # Show first 10 errors
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f'... and {len(errors) - 10} more errors (check server logs)')
        
        if created_count == 0 and updated_count == 0 and error_count == 0 and skipped_count > 0:
            messages.warning(request, '⚠️ No properties were imported. All properties already exist. Check "Force Replace" to update them.')
        
        if created_count == 0 and updated_count == 0 and error_count == 0 and skipped_count == 0:
            messages.warning(request, '⚠️ No valid data found in the Excel file.')
        
    except Exception as e:
        messages.error(request, f'❌ Import failed: {str(e)}')
        print(f"❌ Import exception: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return redirect('excel_import_export')

@login_required
def excel_export(request):
    """Export filtered properties to Excel"""
    if not request.user.is_superuser:
        messages.error(request, 'Access denied: Admins only')
        return redirect('property_list')

    # ⭐ چک کردن template mode
    is_template = request.GET.get('template', '').lower() == 'true'

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties"

    # Define headers
    headers = [
        'Number', 'Team', 'Address ID', 'Village', 'Street', 'House number', 'House number affix',
        'Owner email', 'Owner name', 'Owner surname', 'Owner phone 1', 'Owner phone 2',
        'PoP code', 'Gebaute Units', 'HBG', 'HBG Termin', 'Ausbau Termin',
        'K.L 15M', 'K.L 20M', 'K.L 30M', 'K.L 50M', 'K.L 80M', 'K.L 100M',
        'keller', 'HÜP', 'spleissen', 'ohne Infra', 'mit Infra', 'Status', 'Comments'
    ]

    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ⭐⭐⭐ فقط اگر template نباشد، داده بگذار
    if is_template:
        # هیچ داده‌ای اضافه نکن - فقط header
        pass
    else:
        # Get filters from request
        search = request.GET.get('search', '').strip()
        team_filter = request.GET.get('team', '')
        status_filter = request.GET.get('status', '')

        # Get properties
        properties = Property.objects.all()

        if search:
            properties = properties.filter(
                Q(number__icontains=search) |
                Q(village__icontains=search) |
                Q(owner_name__icontains=search) |
                Q(owner_surname__icontains=search)
            )

        if team_filter:
            properties = properties.filter(teams__id=team_filter)

        if status_filter:
            properties = properties.filter(status=status_filter)

        # Add data rows
        for row_num, prop in enumerate(properties, 2):
            ws.cell(row=row_num, column=1, value=prop.number)
            ws.cell(row=row_num, column=2, value=prop.get_team_names())
            ws.cell(row=row_num, column=2, value=prop.address_id)
            ws.cell(row=row_num, column=3, value=prop.village)
            ws.cell(row=row_num, column=4, value=prop.street)
            ws.cell(row=row_num, column=5, value=prop.house_number)
            ws.cell(row=row_num, column=6, value=prop.house_number_affix)
            ws.cell(row=row_num, column=7, value=prop.owner_email)
            ws.cell(row=row_num, column=8, value=prop.owner_name)
            ws.cell(row=row_num, column=9, value=prop.owner_surname)
            ws.cell(row=row_num, column=10, value=prop.owner_phone_1)
            ws.cell(row=row_num, column=11, value=prop.owner_phone_2)
            ws.cell(row=row_num, column=12, value=prop.pop_code)
            ws.cell(row=row_num, column=13, value=prop.gebaute_units)
            ws.cell(row=row_num, column=14, value=prop.hbg)
            ws.cell(row=row_num, column=15, value=prop.hbg_termin.strftime('%Y-%m-%d %H:%M') if prop.hbg_termin else '')
            ws.cell(row=row_num, column=16, value=prop.ausbau_termin.strftime('%Y-%m-%d %H:%M') if prop.ausbau_termin else '')
            ws.cell(row=row_num, column=17, value=prop.kl_15m)
            ws.cell(row=row_num, column=18, value=prop.kl_20m)
            ws.cell(row=row_num, column=19, value=prop.kl_30m)
            ws.cell(row=row_num, column=20, value=prop.kl_50m)
            ws.cell(row=row_num, column=21, value=prop.kl_80m)
            ws.cell(row=row_num, column=22, value=prop.kl_100m)
            ws.cell(row=row_num, column=23, value=prop.keller)
            ws.cell(row=row_num, column=24, value=prop.huep)
            ws.cell(row=row_num, column=25, value=prop.spleissen)
            ws.cell(row=row_num, column=26, value='X' if prop.infra_type == 'ohne' else '')
            ws.cell(row=row_num, column=27, value='X' if prop.infra_type == 'mit' else '')
            ws.cell(row=row_num, column=28, value=prop.get_status_display())
            ws.cell(row=row_num, column=29, value=prop.comments)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Set filename
    if is_template:
        filename = 'techniknet_template.xlsx'
    else:
        filename = f'techniknet_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response
